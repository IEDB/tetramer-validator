from tetramer_validator.validate import validate
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.comments import Comment
import csv

var_names = {
    "pep_seq": "Peptide Sequence",
    "mhc_name": "MHC Molecule",
    "mod_type": "Modification Type",
    "mod_pos": "Modification Position",
}


def parse_excel_file(filename):
    wb = load_workbook(filename)
    ws = wb.active
    messages = []
    header = {
        "Peptide Sequence": -1,
        "Modification Type": -1,
        "Modification Position": -1,
        "MHC Molecule": -1,
    }
    for entry in ws[1]:
        if entry.value in header.keys():
            header[entry.value] = entry.column - 1

    incorrect_header_string = "IncorrectHeader"
    for (key, value) in header.items():
        if value == -1:
            messages.append(
                {
                    "level": "error",
                    "rule": incorrect_header_string + key,
                    "value": value,
                    "field": key,
                    "instructions": f"{key} field is missing. Please add {key} column in header.",
                    "fix": None,
                    "cell": "A1",
                }
            )
    if messages:
        return (messages, True)
    rows = ws.iter_rows(min_row=2)
    any_errors = False

    for row in rows:
        message = validate(
            pep_seq=row[header["Peptide Sequence"]].value,
            mod_type=row[header["Modification Type"]].value,
            mod_pos=row[header["Modification Position"]].value,
            mhc_name=row[header["MHC Molecule"]].value,
        )
        if message:
            list(
                map(
                    lambda error: error.update(
                        {"cell": row[header[var_names[error["field"]]]].coordinate}
                    ),
                    message,
                )
            )
            messages.extend(message)
            any_errors = True
    return (messages, any_errors)


def parse_csv_tsv(filename, delimiter):
    any_errors = False
    with open(filename, "r", encoding="utf-8-sig") as file_obj:
        reader = csv.DictReader(file_obj, delimiter=delimiter)
        messages = []
        entry_num = 1
        for entry in reader:
            message = validate(
                pep_seq=entry["Peptide Sequence"],
                mhc_name=entry["MHC Molecule"],
                mod_type=entry["Modification Type"],
                mod_pos=entry["Modification Position"],
            )
            if message:
                list(map(lambda error: error.update({"cell": entry_num}), message))
                messages.extend(message)
                any_errors = True
            entry_num += 1
    return (messages, any_errors)


def generate_formatted_data(data_path, problems):
    wb = load_workbook(data_path)
    ws = wb.active
    for problem in problems:
        cell = ws[problem["cell"]]
        if problem["level"] == "error":
            cell.fill = PatternFill(
                patternType="lightUp", fgColor=Color(indexed=10), fill_type="solid"
            )
        else:
            cell.fill = PatternFill(
                patternType="lightUp", fgColor=Color(indexed=52), fill_type="solid"
            )
        if cell.comment:
            comment = Comment(
                cell.comment.text + "\n" + problem["message"],
                author="tetramer_validator",
            )
            cell.comment = comment
        else:
            cell.comment = Comment(problem["message"], author="tetramer-validator")

    wb.save(data_path)


def generate_messages_txt(messages, file_obj):
    writer = csv.DictWriter(
        f=file_obj,
        fieldnames=[
            "level",
            "rule",
            "value",
            "field",
            "message",
            "suggestion",
            "cell",
        ],
    )
    writer.writeheader()
    writer.writerows(messages)
    file_obj.close()
